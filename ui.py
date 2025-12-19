import tkinter as tk
from tkinter import ttk, Tk, filedialog, Button, Label, Text, Scrollbar, END, Y, RIGHT, LEFT, BOTH
from unittest import case
from tkintertable import TableCanvas
import warnings
import sys
import os
import pandas as pd
import ast
import datetime
sys.path.append(os.path.join(os.path.dirname(__file__), '../nova_final_stage1'))
import amazon_inventory
import shipmonk_order
import shipmonk_inventory
import shopify
import csv
import os
from datetime import datetime
import threading
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from datetime import datetime

class UI_test:

    def get_sku_mapping(self):
        """返回完整的 SKU 標準化映射表"""
        sku_mapping = {
            # R1 前綴的 SKU -> 標準 SKU
            'R1-SK010BK-5': 'SK010BK-5',
            'R1-SK010GR-5': 'SK010GR-5',
            'R1-SK010BK': 'SK010BK',
            'R1-SK010GR': 'SK010GR',
            'R1-SK016B': 'SK016B',
            'R1-SK016B-C1': 'SK016B-C1',
            'R1-SK016P': 'SK016P',
            'R1-SK020P': 'SK020P',
            'R1-SK020': 'SK020',
            'R1-SK022': 'SK022',
            'R1-SK025': 'SK025',
            'R1-SK025-3': 'SK025-3',
            'R1-SK026': 'SK026',
            'R1-SK026-3': 'SK026-3',
            'R1-SK029': 'SK029',
            'R1-SB009P-10': 'SB009P-10',
            'R1-SB010-50': 'SB010-50',
            'R1-SB-011-120': 'SB-011-120',
            'R1-DT-009': 'DT-009',
            'R1-SK028A': 'SK028A',
            'R1-SK023': 'SK023',
            'R1-SK022P': 'SK022P',
            'R1-SK031': 'SK031',
            'R1-SB012-100': 'SB012-100',
            'R1-SK008BG-2': 'SK008BG-2',
            
            # R2 前綴的 SKU -> 標準 SKU
            'R2-SK010BK': 'SK010BK',
            'R2-SK010GR': 'SK010GR',
            'R2-SK016B-C1': 'SK016B-C1',
            'R2-SK020P': 'SK020P',
            'R2-SK022': 'SK022',
            'R2-SK022P': 'SK022P',
            'R2-SK023': 'SK023',
            'R2-SK025': 'SK025',
            'R2-SK025-3': 'SK025-3',
            'R2-SK026': 'SK026',
            'R2-SK026-3': 'SK026-3',
            'R2-SK029': 'SK029',
            'R2-SK031': 'SK031',
            'R2-SB012-100': 'SB012-100',
            'R2-SK010BK-BU': 'SK010BK-BU',
            'R2-SK022-BU': 'SK022-BU',
            'R2-SK022P-BU': 'SK022P-BU',
            'R2-SK023-BU': 'SK023-BU',
            'R2-SK025-3-BU': 'SK025-3-BU',
            'R2-SB012-100-BU': 'SB012-100-BU',
            'R2-SK010M': 'SK010M',
            'R2-SK023M': 'SK023M',
            'R2-SK025M': 'SK025M',
            'R2-SK026M': 'SK026M',
            'R2-SK028A': 'SK028A',
            'R2-SK030': 'SK030',
            'R2-SK023A': 'SK023A',
            'R2-SK008BG-2': 'SK008BG-2',
            'R2-SK009GW': 'SK009GW',
            'R2-SB013-10': 'SB013-10',
            'R2-SK012': 'SK012',
            'R2-SK012A': 'SK012A',
            'R2-SK020-Bulk': 'SK020-Bulk',
            'R2-SK017': 'SK017',
            'R2-SK016B-C1-Fulfullment': 'SK016B-C1',
            'R2-SK020-Fufillment': 'SK020',
            'R2-SB009B-10': 'SB009B-10',
            'R2-SB010-100': 'SB010-100',
            
            # 標準 SKU（保持不變）
            'SK010B-10': 'SK010B-10',
            'SK010G-10': 'SK010G-10',
            'SK020P': 'SK020P',
            'SK022P': 'SK022P',
            'SK025': 'SK025',
            'SK026': 'SK026',
            'SK028': 'SK028',
            'PH-101': 'PH-101',
            'STPS-111': 'STPS-111',
            'DT-009': 'DT-009',
            'SB010-100': 'SB010-100',
            'SB011-120': 'SB011-120',
            'SK025-100': 'SK025-100',
            'SK026-100': 'SK026-100',
            'SK020B': 'SK020B',
            'SK025-3': 'SK025-3',
            'SK029': 'SK029',
            'SB009P-10': 'SB009P-10',
            'SB010-50': 'SB010-50',
            'SK016P': 'SK016P',
            'SK016B-C1': 'SK016B-C1',
            'SK016B': 'SK016B',
            'SK010GR': 'SK010GR',
            'SK010BK': 'SK010BK',
            'SK010GR-5': 'SK010GR-5',
            'SK010BK-5': 'SK010BK-5',
            'SK008BG-2': 'SK008BG-2',
            'SK017': 'SK017',
            'SK028A': 'SK028A',
            'SB-011-120': 'SB-011-120',
            'SK031': 'SK031',
            'FK007': 'FK007',
            'SB009B-10': 'SB009B-10',
            'SK022B': 'SK022B',
            'SK021': 'SK021',
            'SK023': 'SK023',
            'SK030': 'SK030',
            'SB012-100': 'SB012-100',
            'FK002-B05': 'FK002-B05',
            'FK006BL': 'FK006BL',
            'FK006GR': 'FK006GR',
            'FK006BK': 'FK006BK',
            'FK008BL': 'FK008BL',
            'FK008GR': 'FK008GR',
            'FK008BK': 'FK008BK',
            'SK012': 'SK012',
            'SK010M': 'SK010M',
            'SK023M': 'SK023M',
            'SP066-B10': 'SP066-B10',
            'SP068-B20': 'SP068-B20',
            'SB018-100': 'SB018-100',
            'SB019-100': 'SB019-100',

            #special-case:
            'SB010-100NEW' : 'SB010-100',
            'SB028A' : 'SB028',
        }
        return sku_mapping

    def __init__(self, root):
        self.root = root
        self.root.title("Nova Excel Viewer")
        warnings.filterwarnings('ignore')

        # Create variables to store file paths
        self.file_paths = [None, None, None]

        # Store processed DataFrames separately
        self.df_product = None
        self.df_single = None
        self.df_carton = None
        self.folder_name = None
        self.df_combined = None


        self.amazon_inven = None

        self.shipmonk_inven = None
        self.shipmonk_order = None
        self.shopify = None

        # Create main container
        self.main_frame = ttk.Frame(root, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        self.top_button_frame = ttk.Frame(self.main_frame)
        self.top_button_frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 10))
        
        self.top_buttons = []

        # Create file selection buttons
        # addbutton = ttk.Button(
        #     self.main_frame, 
        #     text="add button", 
        #     command=self.add_a_button
        # )
        # addbutton.pack(anchor=tk.E, pady=(5, 0))

        self.display_btn = ttk.Button(
            self.main_frame, 
            text="display button", 
            command=self.excute_api
        )
        self.display_btn.pack(anchor=tk.E, pady=(5, 0))

        # 在初始化時就創建關閉按鈕（但先隱藏）
        self.close_btn = ttk.Button(
            self.main_frame,
            text="關閉窗口",
            command=self.close_application
        )
        # 先不pack，等到需要時再顯示
        self.close_btn_created = False

        #self.top_buttons = [addbutton]

        # ADD THIS LINE - Create processing log area
        log_frame = ttk.Frame(self.main_frame)
        log_frame.pack(fill=tk.X, pady=10)
        
        # Label for the log
        log_label = ttk.Label(log_frame, text="Processing Log", font=('Arial', 11, 'bold'))
        log_label.pack(anchor=tk.W, pady=(0, 5))
        
        # Create frame for text widget and scrollbar
        text_frame = ttk.Frame(log_frame)
        text_frame.pack(fill=tk.X, expand=True)
        
        # Create scrollbar
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Create text widget for processing log
        self.processing_log = tk.Text(
            text_frame,
            wrap=tk.WORD,
            yscrollcommand=scrollbar.set,
            font=('Courier', 9),
            height=8,
            bg='#f5f5f5',
            relief=tk.SUNKEN,
            borderwidth=1
        )
        self.processing_log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Configure scrollbar
        scrollbar.config(command=self.processing_log.yview)
        
        # Make the text widget read-only but allow programmatic insertion
        self.processing_log.config(state=tk.DISABLED)
        
        # Add a clear button
        # clear_btn = ttk.Button(
        #     log_frame,
        #     text="Clear Log",
        #     command=self.clear_processing_log
        # )
        # clear_btn.pack(anchor=tk.E, pady=(5, 0))

        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True) 



    def create_tab(self, index):
        """Create a tab with text widget and scrollbar"""
        tab = ttk.Frame(self.notebook)
        self.tabs.append(tab)
        self.notebook.add(tab, text=f"File {index+1}")
        
        text_frame = ttk.Frame(tab)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        text = tk.Text(
            text_frame, 
            wrap=tk.NONE, 
            yscrollcommand=scrollbar.set,
            font=('Courier', 10)  # Monospaced font for better alignment
        )
        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar.config(command=text.yview)
        self.text_widgets.append(text)

    def log_message(self, message):
        """Add a message to the processing log"""
        self.processing_log.config(state=tk.NORMAL)
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.processing_log.insert(tk.END, f"[{timestamp}] {message}\n")
        self.processing_log.config(state=tk.DISABLED)
        self.processing_log.see(tk.END)  # Auto-scroll to bottom

    def clear_processing_log(self):
        """Clear the processing log"""
        self.processing_log.config(state=tk.NORMAL)
        self.processing_log.delete(1.0, tk.END)
        self.processing_log.config(state=tk.DISABLED)

    def display_data_withheader(self, df, filepath, tab_name=None):
        """顯示 DataFrame 數據（簡化版，包含滾動條）"""
        if tab_name is None:
            filename = os.path.basename(filepath)
            tab_name = os.path.splitext(filename)[0]
        
        # 創建或獲取標籤頁
        if not hasattr(self, f'tab_{tab_name}'):
            tab = ttk.Frame(self.notebook)
            self.notebook.add(tab, text=tab_name)
            setattr(self, f'tab_{tab_name}', tab)
        else:
            tab = getattr(self, f'tab_{tab_name}')
            # 清除舊內容
            for widget in tab.winfo_children():
                widget.destroy()
        
        # 創建 Treeview 和滾動條
        tree = ttk.Treeview(tab)
        
        # 垂直滾動條
        v_scrollbar = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=v_scrollbar.set)
        
        # 水平滾動條
        h_scrollbar = ttk.Scrollbar(tab, orient="horizontal", command=tree.xview)
        tree.configure(xscrollcommand=h_scrollbar.set)
        
        # 設置列
        headers = list(df.columns)
        tree["columns"] = headers
        tree["show"] = "headings"
        
        for col in headers:
            tree.heading(col, text=col, anchor='center')
            tree.column(col, anchor='center', width=100)
        
        # 添加數據
        for _, row in df.iterrows():
            tree.insert("", tk.END, values=row.tolist())
        
        # 佈局
        tree.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        
        self.notebook.select(tab)

    def add_a_button(self):
        # 創建文件選擇按鈕
        self.btn4 = ttk.Button(
            self.top_button_frame, 
            text=f"Select单品库存导出", 
            command=lambda idx=0: self.select_file(idx),
            state="disabled"  # 初始狀態為禁用
        )
        self.btn4.grid(row=0, column=0, padx=5, sticky=tk.W)

        self.btn5 = ttk.Button(
            self.top_button_frame, 
            text=f"Select 产品库存导出", 
            command=lambda idx=1: self.select_file(idx),
            state="disabled"  # 初始狀態為禁用
        )
        self.btn5.grid(row=0, column=1, padx=5, sticky=tk.W)

        self.btn6 = ttk.Button(
            self.top_button_frame, 
            text=f"Select Nova Carton qty", 
            command=lambda idx=2: self.select_file(idx),
            state="disabled"  # 初始狀態為禁用
        )
        self.btn6.grid(row=0, column=2, padx=5, sticky=tk.W)

        # 新增 Amazon Order 按鈕
        self.btn_amazon_4month = ttk.Button(
            self.top_button_frame, 
            text=f"Select Amazon Order 4 Month", 
            command=lambda: self.select_amazon_order_file('4month'),
            state="disabled"
        )
        self.btn_amazon_4month.grid(row=0, column=3, padx=5, sticky=tk.W)

        self.btn_amazon_1month = ttk.Button(
            self.top_button_frame, 
            text=f"Select Amazon Order 1 Month", 
            command=lambda: self.select_amazon_order_file('1month'),
            state="disabled"
        )
        self.btn_amazon_1month.grid(row=0, column=4, padx=5, sticky=tk.W)

        # 文件標籤
        for i in range(3):
            label = ttk.Label(self.top_button_frame, text="No file selected", width=30)
            label.grid(row=1, column=i, padx=5, sticky=tk.W)
            setattr(self, f"file_label_{i}", label)
        
        # Amazon Order 文件標籤
        self.amazon_4month_label = ttk.Label(self.top_button_frame, text="No file selected", width=30)
        self.amazon_4month_label.grid(row=1, column=3, padx=5, sticky=tk.W)
        
        self.amazon_1month_label = ttk.Label(self.top_button_frame, text="No file selected", width=30)
        self.amazon_1month_label.grid(row=1, column=4, padx=5, sticky=tk.W)
        
        # 顯示所有文件按鈕（初始狀態為禁用）
        self.display_btn_files = ttk.Button(
            self.top_button_frame,
            text="Display All Files",
            command=self.display_all_files,
            state="disabled"  # 初始狀態為禁用
        )
        self.display_btn_files.grid(row=2, column=0, columnspan=5, pady=10)
        
        # 存儲按鈕引用
        self.top_buttons.extend([
            self.btn4, self.btn5, self.btn6, 
            self.btn_amazon_4month, self.btn_amazon_1month, 
            self.display_btn_files
        ])
        
        # 存儲 Amazon Order 文件路徑
        self.amazon_order_files = {
            '4month': None,
            '1month': None
        }
        
        self.log_message("文件選擇功能已就緒")

    def select_file(self, index, filetypes=None):
        """Handle file selection"""
        if filetypes is None:
            filetypes = [("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]

        filepath = filedialog.askopenfilename(
            title=f"Select Excel File {index+1}",
            filetypes=filetypes,
            initialdir=os.getcwd())

        if filepath:
            self.file_paths[index] = filepath
            label = getattr(self, f"file_label_{index}")
            label.config(text=os.path.basename(filepath))


    def display_api(self):
        relative_path = self.folder_name
        folder_path = os.path.abspath(relative_path)
        file_paths = []

        # 检查文件夹是否存在
        if not os.path.exists(folder_path):
            self.log_message(f"错误: 文件夹不存在 - {folder_path}")
            return

        # 收集所有文件路径
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                file_paths.append(file_path)

        self.log_message(f"找到 {len(file_paths)} 个文件")

        # 先收集所有 Amazon Order 文件
        amazon_order_files = []
        
        for i, filepath in enumerate(file_paths):
            try:
                filename = os.path.basename(filepath)
                filename_no_ext, ext = os.path.splitext(filename)
                ext = ext.lower()

                def read_file_auto(filepath, **kwargs):
                    """自动根据文件扩展名选择读取方法"""
                    if ext in ['.xlsx', '.xls']:
                        # Excel 文件
                        if ext == '.xlsx':
                            return pd.read_excel(filepath, engine='openpyxl', **kwargs)
                        else:
                            return pd.read_excel(filepath, **kwargs)
                    elif ext == '.csv':
                        # CSV 文件
                        return pd.read_csv(filepath, **kwargs)
                    else:
                        raise ValueError(f"不支持的文件格式: {ext}")
                    
                display_df = None
                tab_name = None

                match filename_no_ext:
                    case name if "shopify" in name.lower():
                        self.log_message(f"处理 Shopify 订单文件: {filename}")
                        df = read_file_auto(filepath, header=0)
                        
                        # 处理 items 列
                        df['items'] = df['items'].apply(ast.literal_eval)
                        df_expanded = df.explode('items')
                        df_expanded['sku'] = df_expanded['items'].apply(lambda x: x['sku'])
                        df_expanded['quantity'] = df_expanded['items'].apply(lambda x: x['quantity'])
                        df_final = df_expanded[['sku','created_at', 'quantity']]
                        
                        # 使用完整的 SKU 處理流程
                        df_final = self.standardize_skus_with_mapping(df_final)
                        df_final = df_final[df_final['sku'].notna() & (df_final['sku'] != '')]
                        
                        # 轉換日期格式
                        df_final['created_at'] = pd.to_datetime(df_final['created_at'])
                        
                        # 計算日期範圍
                        latest_date = df_final['created_at'].max()
                        last_30_days_start = latest_date - pd.Timedelta(days=30)
                        last_4_months_start = latest_date - pd.Timedelta(days=120)
                        
                        self.log_message(f"Shopify數據時間範圍: {df_final['created_at'].min()} 到 {latest_date}")
                        self.log_message(f"最近30天範圍: {last_30_days_start} 到 {latest_date}")
                        self.log_message(f"最近4個月範圍: {last_4_months_start} 到 {latest_date}")
                        
                        # 創建時間標記
                        df_final['is_last_30_days'] = df_final['created_at'] >= last_30_days_start
                        df_final['is_last_4_months'] = df_final['created_at'] >= last_4_months_start
                        
                        # 按時間範圍聚合
                        result_data = []
                        
                        for sku in df_final['sku'].unique():
                            sku_data = df_final[df_final['sku'] == sku]
                            
                            # 最近30天數量
                            quantity_30_days = sku_data[sku_data['is_last_30_days']]['quantity'].sum()
                            
                            # 最近4個月總數量
                            quantity_4_months_total = sku_data[sku_data['is_last_4_months']]['quantity'].sum()
                            
                            # 4個月平均值
                            monthly_avg_4_months = round(quantity_4_months_total / 4, 2)
                            
                            result_data.append({
                                'sku': sku,
                                'shopify_quantity_30_days': quantity_30_days,
                                'shopify_quantity_4_months_total': quantity_4_months_total,
                                'shopify_monthly_avg_4_months': monthly_avg_4_months
                            })
                        
                        df_aggregated = pd.DataFrame(result_data)
                        
                        display_df = df_aggregated
                        self.shopify = df_aggregated
                        tab_name = "Shopify订单"
                        
                        self.log_message(f"Shopify訂單聚合完成: {len(df_aggregated)} 個SKU, 總共 {len(df_final)} 筆訂單項目")


                    case name if "shipmonk_inventory" in name.lower():
                        self.log_message(f"处理 Shipmonk 库存文件: {filename}")
                        df = read_file_auto(filepath, header=0)
                        df_final = df[['sku','total_available']]
                        df_final = df_final[df_final['sku'].notna() & (df_final['sku'] != '')]
                        
                        # 使用完整的 SKU 處理流程
                        df_final = self.standardize_skus_with_mapping(df_final)
                        
                        display_df = df_final
                        self.shipmonk_inven = df_final
                        tab_name = "Shipmonk库存"

                    case name if "shimponk_order" in name.lower():
                        self.log_message(f"处理 Shipmonk 订单文件: {filename}")
                        df = read_file_auto(filepath, header=0)
                        df_final = df[['sku', 'ordered_at', 'quantity']]
                        df_final = df_final[df_final['sku'].notna() & (df_final['sku'] != '')]
                        
                        # 使用完整的 SKU 處理流程
                        df_final = self.standardize_skus_with_mapping(df_final)
                        
                        # 轉換日期格式
                        df_final['ordered_at'] = pd.to_datetime(df_final['ordered_at'])
                        
                        # 計算日期範圍
                        latest_date = df_final['ordered_at'].max()
                        
                        # 最近30天
                        last_30_days_start = latest_date - pd.Timedelta(days=30)
                        df_last_30_days = df_final[df_final['ordered_at'] >= last_30_days_start]
                        
                        # 最近4個月
                        last_4_months_start = latest_date - pd.Timedelta(days=120)
                        df_last_4_months = df_final[df_final['ordered_at'] >= last_4_months_start]
                        
                        # 聚合最近30天數據
                        df_30_days_agg = df_last_30_days.groupby('sku', as_index=False)['quantity'].sum()
                        df_30_days_agg.columns = ['sku', 'shipmonk_quantity_30_days']
                        
                        # 聚合最近4個月數據
                        df_4_months_agg = df_last_4_months.groupby('sku', as_index=False)['quantity'].sum()
                        df_4_months_agg.columns = ['sku', 'shipmonk_quantity_4_months_total']
                        
                        # 計算4個月平均值
                        df_4_months_agg['shipmonk_monthly_avg_4_months'] = round(df_4_months_agg['shipmonk_quantity_4_months_total'] / 4, 2)
                        
                        # 合併所有數據
                        df_aggregated = pd.merge(df_30_days_agg, df_4_months_agg, on='sku', how='outer')
                        
                        # 填充 NaN 值為 0
                        df_aggregated = df_aggregated.fillna(0)
                        
                        # 只保留需要的欄位
                        df_aggregated = df_aggregated[['sku', 'shipmonk_quantity_30_days', 'shipmonk_quantity_4_months_total', 'shipmonk_monthly_avg_4_months']]
                        
                        display_df = df_aggregated
                        self.shipmonk_order = df_aggregated
                        tab_name = "Shipmonk订单"
                        
                        self.log_message(f"Shipmonk訂單聚合完成: 最近30天 {len(df_last_30_days)} 筆訂單, 最近4個月 {len(df_last_4_months)} 筆訂單")

                    case name if "amazon_inventory" in name.lower():
                        self.log_message(f"处理 Amazon 库存文件: {filename}")
                        df = read_file_auto(filepath, header=0)
                        df_final = df[['sellerSku', 'fulfillableQuantity','totalReservedQuantity']]
                        df_final = df_final[df_final['sellerSku'].notna() & (df_final['sellerSku'] != '')]
                        df_final = df_final.rename(columns={'sellerSku': 'sku'})
                        df_final = self.standardize_skus_with_mapping(df_final)
                        display_df = df_final
                        self.amazon_inven = df_final
                        tab_name = "Amazon库存"
                        
                    case name if "amazon_order" in name.lower() or "amazon" in name.lower():
                        # 收集 Amazon Order 文件，稍後統一處理
                        amazon_order_files.append(filepath)
                        self.log_message(f"找到 Amazon Order 文件: {filename}")
                        continue

                    case _:
                        self.log_message(f"跳过未识别的文件: {filename}")
                        continue

                # 显示处理后的数据
                if display_df is not None:
                    self.log_message(f"显示 {filename} - {len(display_df)} 行数据")
                    self.display_data_withheader(display_df, filepath, tab_name)
                else:
                    self.log_message(f"警告: {filename} 没有生成显示数据")
                    
            except Exception as e:
                error_msg = f"处理文件 {filename} 时出错: {str(e)}"
                self.log_message(error_msg)
                print(f"Error: {error_msg}")
                import traceback
                traceback.print_exc()

        if amazon_order_files:
            self.process_amazon_order_files(amazon_order_files)

        self.log_message("所有文件处理完成")
        self.enable_file_buttons()

    def standardize_skus_with_mapping(self, df):
        """使用映射表標準化 SKU 名稱"""
        sku_mapping = self.get_sku_mapping()
        df_std = df.copy()
        

        
        # 使用映射表進行標準化
        df_std['sku'] = df_std['sku'].map(sku_mapping)
        
        # 只保留在映射表中找到的 SKU（移除 NaN 值）
        df_std = df_std[df_std['sku'].notna()]
        

        
        return df_std

    def process_sku_data(self, df):
        
        # 步驟 2: 使用映射表標準化 SKU
        df_processed = self.standardize_skus_with_mapping(df_processed)
        
        return df_processed
    
    def enable_file_buttons(self):
        """啟用所有文件相關按鈕"""
        buttons_to_enable = [
        self.btn4, self.btn5, self.btn6, 
        self.btn_amazon_4month, self.btn_amazon_1month, 
        self.display_btn_files
        ]
        
        for button in buttons_to_enable:
            if button:
                button.config(state="normal")
        
        self.log_message("文件選擇按鈕已啟用")


    def display_all_files(self):
        """處理所有上傳的文件並顯示"""
        # 處理三個 Excel 文件
        for i, filepath in enumerate(self.file_paths):
            if filepath:
                try:
                    filename = os.path.basename(filepath)
                    filename_no_ext, ext = os.path.splitext(filename)
                    ext = ext.lower()

                    def read_excel_auto(filepath, **kwargs):
                        if ext == '.xlsx':
                            return pd.read_excel(filepath, engine='openpyxl', **kwargs)
                        else:
                            return pd.read_excel(filepath, **kwargs)
                    
                    display_df = None
                    tab_name = None

                    match filename_no_ext:
                        case name if "产品库存导出" in name:
                            df = read_excel_auto(filepath, header=0)
                            # 只保留列: 仓库, 产品/SKU, 可用库存
                            col_map = {col: col for col in df.columns}
                            needed = ['仓库', '产品/SKU', '可用库存']
                            missing = [col for col in needed if col not in col_map]
                            if missing:
                                raise KeyError(f"Missing columns in 产品库存导出: {missing}")
                            filtered = df[needed]
                            # 只保留 仓库 == '安大略仓库'
                            filtered = filtered[filtered['仓库'] == '安大略仓库']
                            # 重新排序列，讓 产品/SKU 在第一個
                            filtered = filtered[['产品/SKU', '仓库', '可用库存']]
                            filtered = filtered.rename(columns={'产品/SKU': 'sku'})
                            filtered = self.standardize_skus_with_mapping(filtered)
                            filtered = filtered.groupby('sku', as_index=False)['可用库存'].sum()
                            self.df_product = filtered.reset_index(drop=True)
                            display_df = self.df_product
                            tab_name = "产品库存导出"

                        case name if "单品库存导出" in name:
                            df = read_excel_auto(filepath, header=0)
                            # 只保留列: 商家编码, 系统库存, 商品品牌
                            col_map = {col: col for col in df.columns}
                            needed = ['商家编码', '系统库存', '商品品牌']
                            missing = [col for col in needed if col not in col_map]
                            if missing:
                                raise KeyError(f"Missing columns in 单品库存导出: {missing}")
                            filtered = df[needed]
                            # 只保留 商品品牌 == 'Nova'
                            filtered = filtered[filtered['商品品牌'] == 'Nova']
                            filtered = filtered.rename(columns={'商家编码': 'sku'})
                            filtered = self.standardize_skus_with_mapping(filtered)
                            filtered = filtered.groupby('sku', as_index=False)['系统库存'].sum()
                            self.df_single = filtered.reset_index(drop=True)
                            display_df = self.df_single
                            tab_name = "单品库存导出"
                        
                        case name if "Nova Carton qty" in name:
                            df = read_excel_auto(filepath, header=None)
                            # 跳過前2行，只保留列 0,2,3
                            filtered = df.iloc[2:, [0, 2, 3]].copy()
                            filtered.columns = ['Item No.', 'QTY / CTN', 'Unit']
                            filtered = filtered.rename(columns={'Item No.': 'sku'})
                            filtered = self.standardize_skus_with_mapping(filtered)
                            self.df_carton = filtered.reset_index(drop=True)
                            display_df = self.df_carton
                            tab_name = "Nova Carton qty"
                            
                    if display_df is not None:
                        self.log_message(f"顯示 {filename} - {len(display_df)} 行數據")
                        self.display_data_withheader(display_df, filepath, tab_name)
                    else:
                        self.log_message(f"警告: {filename} 沒有生成顯示數據")
                            
                except Exception as e:
                    self.log_message(f"讀取 {filename} 時出錯: {str(e)}")
            else:
                self.log_message(f"文件 {i+1} 未選擇")
        
        # 處理 Amazon Order 文件
        self.process_amazon_order_files_from_upload()
        
        # 在所有文件顯示完成後，添加合併按鈕
        self.add_merge_button()
        self.remove_file_selection_widgets()


    def remove_file_selection_widgets(self):
        """移除文件選擇相關的所有控件"""
        # 移除按鈕
        buttons_to_remove = [
            'btn4', 'btn5', 'btn6', 'btn_amazon_4month', 'btn_amazon_1month',
            'display_btn_files', 'display_btn'
        ]
        
        for btn_name in buttons_to_remove:
            if hasattr(self, btn_name) and getattr(self, btn_name):
                getattr(self, btn_name).destroy()
                setattr(self, btn_name, None)

        # 移除標籤
        for i in range(3):
            label_name = f"file_label_{i}"
            if hasattr(self, label_name):
                label = getattr(self, label_name)
                if label:
                    label.destroy()
                setattr(self, label_name, None)

        # 移除 Amazon Order 標籤
        if hasattr(self, 'amazon_4month_label'):
            self.amazon_4month_label.destroy()
            self.amazon_4month_label = None
            
        if hasattr(self, 'amazon_1month_label'):
            self.amazon_1month_label.destroy()
            self.amazon_1month_label = None

        self.log_message("文件選擇控件已移除")

    def run_apiandexport_csv(self):
        try:
            # Create timestamp for folder name
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.folder_name = f"csv_data_{timestamp}"
            
            # Create directory
            os.makedirs(self.folder_name, exist_ok=True)
            self.log_message(f"Created directory: {self.folder_name}")
            
            # 運行其他 API（除了 Amazon Order）
            self.log_message("Starting: Amazon Inventory")
            result = amazon_inventory.amazon_inventory()
            if result:
                csv_path = os.path.join(self.folder_name, 'amazon_inventory.csv')
                with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                    fieldnames = ['sellerSku', 'fulfillableQuantity', 'totalReservedQuantity', 'lastUpdatedTime']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(result)
                self.log_message(f"CSV file created with {len(result)} records at {csv_path}")

            self.log_message("Starting: Shipmonk Inventory")
            result = shipmonk_inventory.shipmonk_inventory()
            if result:
                csv_path = os.path.join(self.folder_name, 'shipmonk_inventory.csv')
                with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                    fieldnames = ['sku', 'created_at', 'updated_at', 'total_available']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(result)
                self.log_message(f"CSV file created with {len(result)} records at {csv_path}")

            self.log_message("Starting: Shipmonk Order")
            result = shipmonk_order.shipmonk_order()
            if result:
                csv_path = os.path.join(self.folder_name, 'shipmonk_order.csv')
                with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                    fieldnames = ['order_number', 'customer_email', 'ordered_at', 'sku','quantity']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(result)
                self.log_message(f"CSV file created with {len(result)} records at {csv_path}")

            self.log_message("Starting: Shopify")
            result = shopify.shopify()
            if result:
                csv_path = os.path.join(self.folder_name, 'shopify.csv')
                with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                    fieldnames = ['order_number', 'customer_name', 'created_at', 'items']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(result)
                self.log_message(f"CSV file created with {len(result)} records at {csv_path}")
            
            # Amazon Order 現在改為手動上傳 CSV 文件
            self.log_message("Amazon Order: 請手動上傳 CSV 文件")
            
            # Show completion alert
            self.log_message(f"API 執行完成!\n文件保存在: {self.folder_name}")

            self.root.after(0, self.display_api)
        except Exception as e:
            self.log_message(f"An error occurred: {str(e)}")            

    def excute_api(self):

        if hasattr(self, 'display_btn') and self.display_btn :
            self.display_btn.destroy()
            self.display_btn = None

        thread = threading.Thread(target=self.run_apiandexport_csv)
        thread.daemon = True
        thread.start()

        self.add_a_button()
        

    def show_combined_result(self):
        """顯示合併結果在新的tab中（使用 pack 佈局）"""
        if not hasattr(self, 'combined_tab'):
            self.combined_tab = ttk.Frame(self.notebook)
            self.notebook.add(self.combined_tab, text="合併所有數據")
        
        # 清除舊內容
        for widget in self.combined_tab.winfo_children():
            widget.destroy()
        
        # 創建主框架
        main_frame = ttk.Frame(self.combined_tab)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 添加統計信息
        stats_label = ttk.Label(
            main_frame, 
            text=f"合併數據: {len(self.df_combined)} 行, {len(self.df_combined.columns)} 列",
            font=('Arial', 10, 'bold')
        )
        stats_label.pack(anchor=tk.W, pady=(0, 5))
        
        # 創建 Treeview 和滾動條的容器
        tree_scroll_frame = ttk.Frame(main_frame)
        tree_scroll_frame.pack(fill=tk.BOTH, expand=True)
        
        # 創建 Treeview
        tree = ttk.Treeview(tree_scroll_frame)
        headers = list(self.df_combined.columns)
        tree["columns"] = headers
        tree["show"] = "headings"
        
        # 設置列寬和標題
        for col in headers:
            tree.heading(col, text=col, anchor='center')
            tree.column(col, anchor='center', width=120)
        
        # 添加數據
        display_df = self.df_combined.head(1000) if len(self.df_combined) > 1000 else self.df_combined
        
        for _, row in display_df.iterrows():
            tree.insert("", tk.END, values=row.tolist())
        
        # 滾動條
        vsb = ttk.Scrollbar(tree_scroll_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tree_scroll_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # 佈局 Treeview 和滾動條
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 如果數據太多顯示提示
        if len(self.df_combined) > 1000:
            info_label = ttk.Label(
                main_frame, 
                text=f"提示: 只顯示前1000行數據 (總共 {len(self.df_combined)} 行)",
                font=('Arial', 9),
                foreground='blue'
            )
            info_label.pack(anchor=tk.W, pady=(5, 0))
        
        self.notebook.select(self.combined_tab)

        if hasattr(self, 'merge_btn') and self.merge_btn.winfo_exists():
            self.merge_btn.destroy()

        self.add_outputbtn()

    

    def add_merge_button(self):
        """添加合併所有數據的按鈕"""

        # 創建合併按鈕
        self.merge_btn = ttk.Button(
            self.top_button_frame,
            text="Merge All Data (3 Excel + 5 CSV)",
            command=self.merge_all_data,
            style='Accent.TButton'  # 可選：使用強調樣式
        )
        self.merge_btn.grid(row=2, column=0, columnspan=4, pady=10)
        
        self.log_message("合併數據按鈕已添加 - 點擊來合併所有文件")

    def merge_all_data(self):
        """合併所有數據：3個上傳的Excel + 5個API生成的DataFrame"""
        try:
            self.log_message("開始合併所有數據...")
            
            # 存儲所有DataFrame
            all_dfs = {}
            
            # 1. 處理3個上傳的Excel文件
            excel_files = {
                '产品库存导出': self.df_product,
                '单品库存导出': self.df_single,
                'Nova Carton qty': self.df_carton
            }
            
            for name, df in excel_files.items():
                if df is not None:
                    all_dfs[name] = df
                    self.log_message(f"✓ 已加載: {name}")
                else:
                    self.log_message(f"⚠ 未找到: {name}")
            
            # 2. 處理API生成的5個DataFrame
            api_data = {
                'Shopify訂單': self.shopify,
                'Shipmonk訂單': self.shipmonk_order,
                'Shipmonk庫存': self.shipmonk_inven,
                'Amazon庫存': self.amazon_inven,
                # Amazon訂單 - 使用合併後的數據
                'Amazon訂單': getattr(self, 'amazon_order_combined', None)
            }
            
            for name, df in api_data.items():
                if df is not None:
                    all_dfs[name] = df
                    self.log_message(f"✓ 已加載: {name}")
                else:
                    self.log_message(f"⚠ 未找到: {name}")
            
            # 3. 檢查是否有足夠的數據進行合併
            if len(all_dfs) == 0:
                self.log_message("❌ 錯誤: 沒有找到任何數據進行合併")
                return
                
            if '单品库存导出' not in all_dfs:
                self.log_message("❌ 錯誤: 缺少單品庫存導出文件（作為主表）")
                return
                
            # 4. 合併所有DataFrame（使用單品庫存導出作為主表）
            merged_df = all_dfs['单品库存导出'].copy()
            merged_df['__norm_sku__'] = merged_df.iloc[:, 0].apply(self.normalize_sku)
            merged_df = merged_df.set_index('__norm_sku__')
            
            # 合併其他DataFrame
            for name, df in all_dfs.items():
                if name != '单品库存导出' and df is not None:
                    try:
                        temp = df.copy()
                        # 找到包含SKU的列
                        sku_col = None
                        for col in temp.columns:
                            if 'sku' in col.lower() or '编码' in col.lower() or col == '产品/SKU' or col == '商家编码':
                                sku_col = col
                                break
                        if sku_col is None and len(temp.columns) > 0:
                            sku_col = temp.columns[0]  # 使用第一列作為默認
                        
                        if sku_col:
                            temp['__norm_sku__'] = temp[sku_col].apply(self.normalize_sku)
                            temp = temp.set_index('__norm_sku__')
                            
                            # 合併前確保沒有重複的列名
                            merged_df = merged_df.merge(
                                temp, 
                                left_index=True, 
                                right_index=True, 
                                how='outer', 
                                suffixes=('', f'_{name}')
                            )
                            self.log_message(f"✓ 已合併: {name}")
                        else:
                            self.log_message(f"⚠ 跳過 {name}: 未找到SKU列")
                            
                    except Exception as e:
                        self.log_message(f"✗ 合併 {name} 時出錯: {str(e)}")
            
            # 清理合併後的DataFrame
            merged_df.reset_index(drop=True, inplace=True)
            merged_df.columns = merged_df.columns.map(str)
            merged_df = merged_df.loc[:, ~merged_df.columns.str.contains('__norm_sku__')]
            
            self.df_combined = merged_df

            self.show_combined_result()
            self.log_message(f"✅ 合併完成! 總計 {len(merged_df)} 行, {len(merged_df.columns)} 列")
            
            # 5. 顯示export btn
            if hasattr(self, 'merge_btn'):
                self.merge_btn.destroy()
        
            self.add_outputbtn()
            
        except Exception as e:
            self.log_message(f"❌ 合併過程出錯: {str(e)}")
            import traceback
            traceback.print_exc()
            self.log_message(f"❌ 合併過程出錯: {str(e)}")
            import traceback
            traceback.print_exc()



    def normalize_sku(self, sku):
        if pd.isna(sku):
            return ''
        return str(sku).replace('-', '').lower().strip()
    
    def add_outputbtn(self):
        self.output_btn = ttk.Button(
            self.top_button_frame,
            text="Output csv",
            command=self.export_to_excel,
            style='Accent.TButton'  # 可選：使用強調樣式
        )
        self.output_btn.grid(row=2, column=0, columnspan=4, pady=10)
        
        self.log_message("合併數據按鈕已添加 - 點擊來合併所有文件")
    
    def export_to_excel(self):
        """Export the specified DataFrame to an Excel file."""
        # 確保合併 DataFrame 可用
        if self.df_combined is None:
            self.log_message("錯誤: 沒有合併數據可供導出")
            return
        
        now = datetime.now()
        filename = f"nova inventory_{now.month:02d}{now.day:02d}_{now.hour:02d}{now.minute:02d}.xlsx"

        wb = Workbook()
        us_sea = wb.active
        us_sea.title = "US Sea"
        air_shipment = wb.create_sheet("Air Shipment")
        new_po = wb.create_sheet("New PO")

        # 設置標題和格式
        us_sea.merge_cells('A1:D1')
        us_sea['A1'].value = "US SKU Sea Shipment"
        us_sea['A1'].font = Font(color="FF156082")

        air_shipment.merge_cells('A1:D1')
        air_shipment['A1'].value = "US SKU Air Shipment"
        air_shipment['A1'].font = Font(color="FF196B24")

        new_po.merge_cells('A1:D1')
        new_po['A1'].value = "Issue PO"
        new_po['A1'].font = Font(color="FFA02B93")

        # 設置表頭
        headers = {
            'us_sea': [
                "SKU", "US Warehouse Qty", "Amazon FBA Qty", "ShipMonk Inventory",
                "China Inventory", "Latest 30 days Amazon sales", "Average 4 month Sales",
                "Latest 30 days Shopify Sales", "Latest 30 days shipmonk Sales",
                "Replenishment Qty (7M monthly sales)", "Unit"
            ],
            'air_shipment': [
                "SKU", "US Warehouse Qty", "Amazon FBA Qty", "ShipMonk Inventory",
                "China Inventory", "Latest 30 days Amazon Sales", "Average 4 month Sales",
                "Latest 30 days Shopify Sales", "Latest 30 days shipmonk Sales",
                "Replenishment Qty (1.5M monthly sales)", "Unit"
            ],
            'new_po': [
                "SKU", "China Inventory", "Reserved qty (From Air Shipment and Sea Shipment)",
                "Suggested Replenishment Qty (From program suggestions)", "Inventory left",
                "Open Order Qty", "Latest 30 days Total sales", "Average 6 month Sales",
                "Order Qty", "Unit"
            ]
        }

        # 應用表頭格式
        sheet_configs = {
            'us_sea': (us_sea, "FF156082", 3),
            'air_shipment': (air_shipment, "FF196B24", 3),
            'new_po': (new_po, "FFA02B93", 3)
        }

        white = Font(color="FFFFFFFF")

        for sheet_name, (sheet, color, row) in sheet_configs.items():
            header_list = headers[sheet_name]
            for col, header in enumerate(header_list, start=1):
                cell = sheet.cell(row=row, column=col)
                cell.value = header
                cell.fill = PatternFill(start_color=color, fill_type="solid")
                cell.font = white

        # 初始化數據行計數器
        us_sea_row = 4
        air_shipment_row = 4
        new_po_row = 4

        # 確保有 SKU 列
        if 'sku' not in self.df_combined.columns:
            # 嘗試找到 SKU 列
            sku_cols = [col for col in self.df_combined.columns if 'sku' in col.lower()]
            if sku_cols:
                sku_col = sku_cols[0]
                self.df_combined = self.df_combined.rename(columns={sku_col: 'sku'})
            else:
                self.log_message("錯誤: 找不到 SKU 列")
                return

        df = self.df_combined.dropna(subset=['sku'])
        self.log_message(f"開始處理 {len(df)} 個SKU 的導出")

        # Group by SKU
        for sku_code, group in df.groupby('sku'):
            if pd.isna(sku_code) or sku_code == '':
                continue

            # Basic info (取第一個非空值)
            china_store = group['系统库存'].dropna().iloc[0] if '系统库存' in group.columns and not group['系统库存'].dropna().empty else 0
            US_store = group['可用库存'].dropna().iloc[0] if '可用库存' in group.columns and not group['可用库存'].dropna().empty else 0
            shipmonk_inventory = group['total_available'].dropna().iloc[0] if 'total_available' in group.columns and not group['total_available'].dropna().empty else 0
            QTY = group['QTY / CTN'].dropna().iloc[0] if 'QTY / CTN' in group.columns and not group['QTY / CTN'].dropna().empty else 1  # 默認為1避免除零錯誤
            Unit = group['Unit'].dropna().iloc[0] if 'Unit' in group.columns and not group['Unit'].dropna().empty else 'PCS'
            Open_order = 0
            Amazon_FBA = group['fulfillableQuantity'].dropna().iloc[0] if 'fulfillableQuantity' in group.columns and not group['fulfillableQuantity'].dropna().empty else 0
            
            # Amazon 數據 - 使用會話數
            Amazon_average = 0
            Amazon_apr = 0
            
            # 嘗試不同列名獲取 Amazon 數據
            amazon_avg_cols = ['amazon_monthly_avg_4_months', 'orders_4month', 'amazon_order_orders_4month']
            amazon_apr_cols = ['amazon_orderItemCount_30_days', 'orders_1month', 'amazon_order_orders_1month']
            
            for col in amazon_avg_cols:
                if col in group.columns and not group[col].dropna().empty:
                    Amazon_average = group[col].dropna().iloc[0] / 4 if col == 'orders_4month' or col == 'amazon_order_orders_4month' else group[col].dropna().iloc[0]
                    break
            
            for col in amazon_apr_cols:
                if col in group.columns and not group[col].dropna().empty:
                    Amazon_apr = group[col].dropna().iloc[0]
                    break
            
            # Shopify 和 Shipmonk 數據
            shopfit_average = group['shopify_monthly_avg_4_months'].dropna().iloc[0] if 'shopify_monthly_avg_4_months' in group.columns and not group['shopify_monthly_avg_4_months'].dropna().empty else 0
            shopfit_apr = group['shopify_quantity_30_days'].dropna().iloc[0] if 'shopify_quantity_30_days' in group.columns and not group['shopify_quantity_30_days'].dropna().empty else 0
            shipmonk_average = group['shipmonk_monthly_avg_4_months'].dropna().iloc[0] if 'shipmonk_monthly_avg_4_months' in group.columns and not group['shipmonk_monthly_avg_4_months'].dropna().empty else 0
            shipmonk_apr = group['shipmonk_quantity_30_days'].dropna().iloc[0] if 'shipmonk_quantity_30_days' in group.columns and not group['shipmonk_quantity_30_days'].dropna().empty else 0
            
            # 初始化補貨數量
            air_shipment_replenishment = 0
            shipping_replenishment = 0
            china_store_updated = china_store

            # 計算 US 倉庫庫存（考慮每箱數量）
            us_warehouse_qty = US_store * QTY if US_store > 0 else 0

            # 調試信息
            self.log_message(f"處理 SKU: {sku_code}")
            self.log_message(f"  Amazon平均: {Amazon_average:.2f}, Amazon最近30天: {Amazon_apr}")
            self.log_message(f"  Shopify平均: {shopfit_average:.2f}, Shopify最近30天: {shopfit_apr}")
            self.log_message(f"  Shipmonk平均: {shipmonk_average:.2f}, Shipmonk最近30天: {shipmonk_apr}")
            self.log_message(f"  中國庫存: {china_store}, US倉庫: {us_warehouse_qty}, Shipmonk庫存: {shipmonk_inventory}, Amazon FBA: {Amazon_FBA}")

            # 計算總銷售數據
            total_30_days = Amazon_apr + shopfit_apr + shipmonk_apr
            total_avg = Amazon_average + shopfit_average + shipmonk_average
            total_inventory = us_warehouse_qty + shipmonk_inventory + Amazon_FBA

            self.log_message(f"  最近30天總銷售: {total_30_days:.2f}, 平均月銷售: {total_avg:.2f}, 總庫存: {total_inventory:.2f}")

            # 空運補貨邏輯：如果最近30天銷售的1.5倍 > 總庫存
            if total_30_days * 1.5 > total_inventory:
                raw_qty = total_avg * 1.5
                box_count = max(1, round(raw_qty / QTY))
                air_shipment_replenishment = box_count * QTY
                
                # 寫入空運工作表
                air_shipment.cell(row=air_shipment_row, column=1).value = sku_code
                air_shipment.cell(row=air_shipment_row, column=2).value = us_warehouse_qty
                air_shipment.cell(row=air_shipment_row, column=3).value = Amazon_FBA
                air_shipment.cell(row=air_shipment_row, column=4).value = shipmonk_inventory
                air_shipment.cell(row=air_shipment_row, column=5).value = china_store
                air_shipment.cell(row=air_shipment_row, column=6).value = Amazon_apr
                air_shipment.cell(row=air_shipment_row, column=7).value = total_avg
                air_shipment.cell(row=air_shipment_row, column=8).value = shopfit_apr
                air_shipment.cell(row=air_shipment_row, column=9).value = shipmonk_apr
                air_shipment.cell(row=air_shipment_row, column=10).value = air_shipment_replenishment
                air_shipment.cell(row=air_shipment_row, column=11).value = Unit
                air_shipment_row += 1
                
                china_store_updated = china_store - air_shipment_replenishment
                self.log_message(f"  空運補貨: {air_shipment_replenishment}, 中國庫存更新後: {china_store_updated}")

            # 海運補貨邏輯：如果最近30天銷售的4倍 > 總庫存 + 空運補貨
            if total_30_days * 4 > (total_inventory + air_shipment_replenishment):
                raw_qty = total_avg * 7
                box_count = max(1, round(raw_qty / QTY))
                shipping_replenishment = box_count * QTY - air_shipment_replenishment
                
                if shipping_replenishment < 0:
                    shipping_replenishment = 0
                
                # 寫入海運工作表
                us_sea.cell(row=us_sea_row, column=1).value = sku_code
                us_sea.cell(row=us_sea_row, column=2).value = us_warehouse_qty
                us_sea.cell(row=us_sea_row, column=3).value = Amazon_FBA
                us_sea.cell(row=us_sea_row, column=4).value = shipmonk_inventory
                us_sea.cell(row=us_sea_row, column=5).value = china_store
                us_sea.cell(row=us_sea_row, column=6).value = Amazon_apr
                us_sea.cell(row=us_sea_row, column=7).value = total_avg
                us_sea.cell(row=us_sea_row, column=8).value = shopfit_apr
                us_sea.cell(row=us_sea_row, column=9).value = shipmonk_apr
                us_sea.cell(row=us_sea_row, column=10).value = shipping_replenishment
                us_sea.cell(row=us_sea_row, column=11).value = Unit
                us_sea_row += 1
                
                china_store_updated -= shipping_replenishment
                self.log_message(f"  海運補貨: {shipping_replenishment}, 中國庫存更新後: {china_store_updated}")

            # 新採購單邏輯：如果剩餘中國庫存 + 開放訂單 < 最近30天銷售的7倍
            if (china_store_updated + Open_order) < (total_30_days * 7):
                raw_qty = total_avg * 14
                box_count = max(1, round(raw_qty / QTY))
                new_po_replenishment = box_count * QTY - Open_order
                
                if new_po_replenishment < 0:
                    new_po_replenishment = QTY  # 至少一箱
                
                # 寫入新採購單工作表
                new_po.cell(row=new_po_row, column=1).value = sku_code
                new_po.cell(row=new_po_row, column=2).value = china_store
                new_po.cell(row=new_po_row, column=3).value = 0
                new_po.cell(row=new_po_row, column=4).value = (air_shipment_replenishment + shipping_replenishment)
                new_po.cell(row=new_po_row, column=5).value = china_store_updated
                new_po.cell(row=new_po_row, column=6).value = Open_order
                new_po.cell(row=new_po_row, column=7).value = total_30_days
                new_po.cell(row=new_po_row, column=8).value = total_avg
                new_po.cell(row=new_po_row, column=9).value = new_po_replenishment
                new_po.cell(row=new_po_row, column=10).value = Unit
                new_po_row += 1
                
                self.log_message(f"  新採購單補貨: {new_po_replenishment}")

            self.log_message("")  # 空行分隔

        # 保存文件
        wb.save(filename)
        self.log_message(f"✅ Excel 文件已導出: {filename}")

        # 移除輸出按鈕，顯示關閉按鈕
        if hasattr(self, 'output_btn') and self.output_btn.winfo_exists():
            self.output_btn.destroy()

        if not self.close_btn_created:
            self.close_btn.pack(anchor=tk.E, pady=(10, 5))
            self.close_btn_created = True

    def close_application(self):
        """關閉應用程序"""
        self.log_message("正在關閉應用程序...")
        self.root.quit()
        self.root.destroy()

    def process_amazon_order_files(self, amazon_order_files):
        """處理 Amazon Order CSV 文件"""
        try:
            self.log_message(f"開始處理 {len(amazon_order_files)} 個 Amazon Order 文件")
            
            # 讀取並處理每個 Amazon Order 文件
            amazon_order_data = []
            
            for filepath in amazon_order_files:
                filename = os.path.basename(filepath)
                self.log_message(f"處理 Amazon Order 文件: {filename}")
                
                # 讀取 CSV 文件
                df = pd.read_csv(filepath)
                
                # 檢查必要的列是否存在
                if 'SKU' not in df.columns or 'Units Ordered' not in df.columns:
                    self.log_message(f"警告: {filename} 缺少必要的列 (SKU 或 Units Ordered)")
                    continue
                
                # 選擇需要的列
                df_processed = df[['SKU', 'Units Ordered']].copy()
                
                # 清理數據
                df_processed = df_processed[df_processed['SKU'].notna() & (df_processed['SKU'] != '')]
                
                # 清理 Units Ordered 列（移除逗號並轉換為數字）
                df_processed['Units Ordered'] = df_processed['Units Ordered'].astype(str).str.replace(',', '').astype(float)
                
                # 標準化 SKU
                df_processed = self.standardize_skus_with_mapping(df_processed.rename(columns={'SKU': 'sku'}))
                
                # 根據文件名判斷是4個月還是1個月的數據
                filename_lower = filename.lower()
                if '4month' in filename_lower or '4_month' in filename_lower:
                    period_type = '4month'
                    df_processed = df_processed.rename(columns={'Units Ordered': 'orders_4month'})
                elif '1month' in filename_lower or '1_month' in filename_lower:
                    period_type = '1month'
                    df_processed = df_processed.rename(columns={'Units Ordered': 'orders_1month'})
                else:
                    # 如果文件名沒有明確標示，根據文件數量判斷
                    if len(amazon_order_files) == 2:
                        period_type = '4month' if len(amazon_order_data) == 0 else '1month'
                        column_name = 'orders_4month' if period_type == '4month' else 'orders_1month'
                        df_processed = df_processed.rename(columns={'Units Ordered': column_name})
                    else:
                        period_type = 'unknown'
                        self.log_message(f"警告: 無法確定 {filename} 的時間週期")
                        continue
                
                amazon_order_data.append((period_type, df_processed))
                self.log_message(f"✓ 已處理 {filename} 作為 {period_type} 數據")
            
            # 合併 Amazon Order 數據
            if len(amazon_order_data) >= 1:
                self.merge_and_display_amazon_orders(amazon_order_data)
            else:
                self.log_message("❌ 沒有有效的 Amazon Order 數據")
                
        except Exception as e:
            self.log_message(f"處理 Amazon Order 文件時出錯: {str(e)}")
            import traceback
            traceback.print_exc()

    def merge_and_display_amazon_orders(self, amazon_order_data):
        """合併並顯示 Amazon Order 數據在同一個表格中"""
        try:
            # 創建合併的 DataFrame
            merged_df = None
            
            for period_type, df in amazon_order_data:
                if merged_df is None:
                    merged_df = df.copy()
                else:
                    # 使用 outer join 合併兩個時期的數據
                    merged_df = pd.merge(merged_df, df, on='sku', how='outer')
            
            # 填充 NaN 值為 0
            merged_df = merged_df.fillna(0)
            
            # 計算平均值和最近30天數據
            if 'orders_4month' in merged_df.columns:
                # 4個月平均數 = 4個月總會話數 / 4
                merged_df['amazon_monthly_avg_4_months'] = merged_df['orders_4month'] / 4
            
            if 'orders_1month' in merged_df.columns:
                # 最近30天數據 = 1個月會話數
                merged_df['amazon_orderItemCount_30_days'] = merged_df['orders_1month']
            
                
            
            # 保存到實例變量
            self.amazon_order_combined = merged_df.copy()
            
            # 分別保存到獨立變量
            if 'orders_4month' in merged_df.columns:
                self.amazon_order_4month = merged_df[['sku', 'orders_4month']].copy()
            if 'orders_1month' in merged_df.columns:
                self.amazon_order_1month = merged_df[['sku', 'orders_1month']].copy()
            
            # 顯示合併後的數據
            self.log_message(f"Amazon Order 合併完成: {len(merged_df)} 個SKU")
            self.log_message(f"Amazon 4個月平均計算完成")
            self.log_message(f"Amazon 最近30天數據計算完成")
            
            # 顯示在 Amazon訂單合併 tab
            self.display_data_withheader(merged_df, "Amazon Orders Combined", "Amazon訂單合併")
            
            return merged_df
            
        except Exception as e:
            self.log_message(f"合併 Amazon Order 數據時出錯: {str(e)}")
            import traceback
            traceback.print_exc()
            return None


    
    def select_amazon_order_file(self, period_type):
        """選擇 Amazon Order 文件"""
        filetypes = [("CSV files", "*.csv"), ("All files", "*.*")]
        
        filepath = filedialog.askopenfilename(
            title=f"Select Amazon Order {period_type} File",
            filetypes=filetypes,
            initialdir=os.getcwd()
        )

        if filepath:
            self.amazon_order_files[period_type] = filepath
            label = getattr(self, f"amazon_{period_type}_label")
            label.config(text=os.path.basename(filepath))
            self.log_message(f"已選擇 Amazon Order {period_type} 文件: {os.path.basename(filepath)}")

    def process_amazon_order_files_from_upload(self):
        """處理上傳的 Amazon Order 文件"""
        try:
            amazon_order_data = []
            
            # 處理 4個月數據
            if self.amazon_order_files['4month']:
                filepath = self.amazon_order_files['4month']
                df = self.read_and_process_amazon_order(filepath, '4month')
                if df is not None:
                    amazon_order_data.append(('4month', df))
            
            # 處理 1個月數據
            if self.amazon_order_files['1month']:
                filepath = self.amazon_order_files['1month']
                df = self.read_and_process_amazon_order(filepath, '1month')
                if df is not None:
                    amazon_order_data.append(('1month', df))
            
            # 合併並顯示 Amazon Order 數據
            if amazon_order_data:
                merged_df = self.merge_and_display_amazon_orders(amazon_order_data)
                if merged_df is not None:
                    # 顯示詳細信息
                    order_columns = [col for col in merged_df.columns if 'orders' in col]
                    self.log_message(f"Amazon Order 數據包含以下會話列: {', '.join(order_columns)}")
                    
                    for period in ['4month', '1month']:
                        if f'orders_{period}' in merged_df.columns:
                            total_orders = merged_df[f'orders_{period}'].sum()
                            self.log_message(f"Amazon Order {period} 總會話數: {total_orders:,.0f}")
            else:
                self.log_message("❌ 沒有有效的 Amazon Order 數據")
                
        except Exception as e:
            self.log_message(f"處理 Amazon Order 文件時出錯: {str(e)}")
            import traceback
            traceback.print_exc()

    def read_and_process_amazon_order(self, filepath, period_type):
        """讀取並處理單個 Amazon Order CSV 文件"""
        try:
            filename = os.path.basename(filepath)
            self.log_message(f"處理 Amazon Order {period_type} 文件: {filename}")
            
            # 讀取 CSV 文件
            df = pd.read_csv(filepath)
            
            # 檢查必要的列是否存在
            if 'SKU' not in df.columns:
                # 嘗試其他可能的列名
                possible_sku_cols = ['SKU', 'Seller SKU', 'sku', 'sellerSku']
                sku_col = None
                for col in possible_sku_cols:
                    if col in df.columns:
                        sku_col = col
                        break
                
                if sku_col is None:
                    self.log_message(f"警告: {filename} 缺少 SKU 列")
                    return None
            else:
                sku_col = 'SKU'
            
            if 'Units Ordered' not in df.columns:
                # 嘗試其他可能的會話列名
                possible_session_cols = ['Units Ordered', 'Sessions', 'sessions', 'Sessions_Total']
                session_col = None
                for col in possible_session_cols:
                    if col in df.columns:
                        session_col = col
                        break
                
                if session_col is None:
                    self.log_message(f"警告: {filename} 缺少會話數據列")
                    return None
            else:
                session_col = 'Units Ordered'
            
            # 選擇需要的列
            df_processed = df[[sku_col, session_col]].copy()
            
            # 清理數據
            df_processed = df_processed[df_processed[sku_col].notna() & (df_processed[sku_col] != '')]
            
            # 重命名列
            df_processed = df_processed.rename(columns={sku_col: 'sku', session_col: 'Units Ordered'})
            
            # 清理 Units Ordered 列（移除逗號、貨幣符號等並轉換為數字）
            df_processed['Units Ordered'] = (
                df_processed['Units Ordered']
                .astype(str)
                .str.replace(',', '')
                .str.replace('$', '')
                .str.replace('US', '')
                .str.replace(' ', '')
                .apply(pd.to_numeric, errors='coerce')
            )
            
            # 標準化 SKU
            df_processed = self.standardize_skus_with_mapping(df_processed)
            
            # 按 SKU 分組求和（如果有多行相同 SKU）
            df_processed = df_processed.groupby('sku', as_index=False)['Units Ordered'].sum()
            
            # 重命名列
            column_name = f'orders_{period_type}'
            df_processed = df_processed.rename(columns={'Units Ordered': column_name})
            
            self.log_message(f"✓ 已處理 {period_type} 數據: {len(df_processed)} 個SKU")
            
            return df_processed
            
        except Exception as e:
            self.log_message(f"讀取 Amazon Order {period_type} 文件時出錯: {str(e)}")
            import traceback
            traceback.print_exc()
            return None



if __name__ == "__main__":
    root = tk.Tk()
    app = UI_test(root)
    root.geometry("1200x800")  # Slightly larger window
    root.mainloop()